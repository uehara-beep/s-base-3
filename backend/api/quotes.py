from fastapi import APIRouter, Depends, HTTPException, File, UploadFile
from sqlalchemy.orm import Session, joinedload
from typing import List, Optional
from pydantic import BaseModel
from datetime import date, datetime
from decimal import Decimal
from database import get_db
from models.quote import Quote, QuoteItem
from models.user import User
from utils.auth import get_current_user, require_role
import openpyxl
from io import BytesIO
import re

router = APIRouter()


class QuoteItemCreate(BaseModel):
    item_order: int = 0
    category: Optional[str] = None
    description: str
    specification: Optional[str] = None
    quantity: float = 1
    unit: Optional[str] = None
    unit_price: float = 0
    cost_price: float = 0
    notes: Optional[str] = None


class QuoteCreate(BaseModel):
    client_id: Optional[int] = None
    project_name: str
    site_name: Optional[str] = None
    site_address: Optional[str] = None
    quote_date: Optional[date] = None
    valid_until: Optional[date] = None
    tax_rate: float = 10
    notes: Optional[str] = None
    items: List[QuoteItemCreate] = []


class QuoteResponse(BaseModel):
    id: int
    quote_number: str
    project_name: str
    site_name: Optional[str]
    quote_date: Optional[date]
    subtotal: float
    tax_amount: float
    total_amount: float
    cost_amount: float
    profit_amount: float
    profit_rate: float
    status: str

    class Config:
        from_attributes = True


@router.get("/", response_model=List[QuoteResponse])
async def get_quotes(
    status: Optional[str] = None,
    client_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    query = db.query(Quote)
    if status:
        query = query.filter(Quote.status == status)
    if client_id:
        query = query.filter(Quote.client_id == client_id)
    return query.order_by(Quote.created_at.desc()).all()


@router.get("/{quote_id}")
async def get_quote(
    quote_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    quote = db.query(Quote).options(joinedload(Quote.items)).filter(Quote.id == quote_id).first()
    if not quote:
        raise HTTPException(status_code=404, detail="見積が見つかりません")

    return {
        "id": quote.id,
        "quote_number": quote.quote_number,
        "client_id": quote.client_id,
        "project_name": quote.project_name,
        "site_name": quote.site_name,
        "site_address": quote.site_address,
        "quote_date": quote.quote_date,
        "valid_until": quote.valid_until,
        "subtotal": float(quote.subtotal or 0),
        "tax_rate": float(quote.tax_rate or 10),
        "tax_amount": float(quote.tax_amount or 0),
        "total_amount": float(quote.total_amount or 0),
        "cost_amount": float(quote.cost_amount or 0),
        "profit_amount": float(quote.profit_amount or 0),
        "profit_rate": float(quote.profit_rate or 0),
        "status": quote.status,
        "notes": quote.notes,
        "items": [
            {
                "id": item.id,
                "item_order": item.item_order,
                "category": item.category,
                "description": item.description,
                "specification": item.specification,
                "quantity": float(item.quantity or 0),
                "unit": item.unit,
                "unit_price": float(item.unit_price or 0),
                "amount": float(item.amount or 0),
                "cost_price": float(item.cost_price or 0),
                "cost_amount": float(item.cost_amount or 0),
            }
            for item in sorted(quote.items, key=lambda x: x.item_order)
        ]
    }


@router.post("/")
async def create_quote(
    data: QuoteCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    # Generate quote number
    last_quote = db.query(Quote).order_by(Quote.id.desc()).first()
    quote_num = f"Q{(last_quote.id + 1 if last_quote else 1):06d}"

    # Calculate totals
    subtotal = Decimal(0)
    cost_total = Decimal(0)

    quote = Quote(
        quote_number=quote_num,
        client_id=data.client_id,
        project_name=data.project_name,
        site_name=data.site_name,
        site_address=data.site_address,
        quote_date=data.quote_date or date.today(),
        valid_until=data.valid_until,
        tax_rate=Decimal(str(data.tax_rate)),
        notes=data.notes,
        created_by=current_user.id
    )

    db.add(quote)
    db.flush()

    # Add items
    for item_data in data.items:
        amount = Decimal(str(item_data.quantity)) * Decimal(str(item_data.unit_price))
        cost_amount = Decimal(str(item_data.quantity)) * Decimal(str(item_data.cost_price))

        item = QuoteItem(
            quote_id=quote.id,
            item_order=item_data.item_order,
            category=item_data.category,
            description=item_data.description,
            specification=item_data.specification,
            quantity=Decimal(str(item_data.quantity)),
            unit=item_data.unit,
            unit_price=Decimal(str(item_data.unit_price)),
            amount=amount,
            cost_price=Decimal(str(item_data.cost_price)),
            cost_amount=cost_amount,
            notes=item_data.notes
        )
        db.add(item)
        subtotal += amount
        cost_total += cost_amount

    # Update quote totals
    quote.subtotal = subtotal
    quote.tax_amount = subtotal * quote.tax_rate / 100
    quote.total_amount = subtotal + quote.tax_amount
    quote.cost_amount = cost_total
    quote.profit_amount = subtotal - cost_total
    quote.profit_rate = (quote.profit_amount / subtotal * 100) if subtotal > 0 else Decimal(0)

    db.commit()
    db.refresh(quote)

    return {"message": "見積が作成されました", "quote_id": quote.id, "quote_number": quote.quote_number}


@router.put("/{quote_id}")
async def update_quote(
    quote_id: int,
    data: QuoteCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    quote = db.query(Quote).filter(Quote.id == quote_id).first()
    if not quote:
        raise HTTPException(status_code=404, detail="見積が見つかりません")

    # Update basic info
    quote.client_id = data.client_id
    quote.project_name = data.project_name
    quote.site_name = data.site_name
    quote.site_address = data.site_address
    quote.quote_date = data.quote_date
    quote.valid_until = data.valid_until
    quote.tax_rate = Decimal(str(data.tax_rate))
    quote.notes = data.notes

    # Delete existing items
    db.query(QuoteItem).filter(QuoteItem.quote_id == quote_id).delete()

    # Add new items
    subtotal = Decimal(0)
    cost_total = Decimal(0)

    for item_data in data.items:
        amount = Decimal(str(item_data.quantity)) * Decimal(str(item_data.unit_price))
        cost_amount = Decimal(str(item_data.quantity)) * Decimal(str(item_data.cost_price))

        item = QuoteItem(
            quote_id=quote.id,
            item_order=item_data.item_order,
            category=item_data.category,
            description=item_data.description,
            specification=item_data.specification,
            quantity=Decimal(str(item_data.quantity)),
            unit=item_data.unit,
            unit_price=Decimal(str(item_data.unit_price)),
            amount=amount,
            cost_price=Decimal(str(item_data.cost_price)),
            cost_amount=cost_amount,
            notes=item_data.notes
        )
        db.add(item)
        subtotal += amount
        cost_total += cost_amount

    # Update totals
    quote.subtotal = subtotal
    quote.tax_amount = subtotal * quote.tax_rate / 100
    quote.total_amount = subtotal + quote.tax_amount
    quote.cost_amount = cost_total
    quote.profit_amount = subtotal - cost_total
    quote.profit_rate = (quote.profit_amount / subtotal * 100) if subtotal > 0 else Decimal(0)

    db.commit()
    return {"message": "見積が更新されました"}


@router.put("/{quote_id}/status")
async def update_quote_status(
    quote_id: int,
    status: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    quote = db.query(Quote).filter(Quote.id == quote_id).first()
    if not quote:
        raise HTTPException(status_code=404, detail="見積が見つかりません")

    quote.status = status
    db.commit()
    return {"message": f"見積のステータスが{status}に変更されました"}


@router.delete("/{quote_id}")
async def delete_quote(
    quote_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(["admin", "manager"]))
):
    quote = db.query(Quote).filter(Quote.id == quote_id).first()
    if not quote:
        raise HTTPException(status_code=404, detail="見積が見つかりません")

    db.delete(quote)
    db.commit()
    return {"message": "見積が削除されました"}


@router.post("/import-excel")
async def import_excel(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user)
):
    """
    Excelファイルから見積データを読み込む（1.0互換方式）
    - 全シートをスキャンして情報を抽出
    - 柔軟なセル解析で様々なフォーマットに対応
    """
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Excelファイル(.xlsx, .xls)のみ対応しています")

    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(BytesIO(contents), data_only=True)

        result = {
            "cover": {},
            "items": [],
            "conditions": [],
            "confirmation": {"items": [], "special_notes": ""}
        }

        sheet_names = wb.sheetnames

        # スキップするキーワード
        skip_keywords = ['小計', '合計', '端数調整', '法定福利費', '諸経費計', '消費税']
        # 単位として認識するキーワード
        unit_keywords = ['式', '個', '本', 'm', 'm2', 'm3', 'kg', 't', 'L', '人工', '台', '日', '往復', 'セット', '箇所', '回']

        # ============ 1. 表紙・基本情報の解析 ============
        for sheet_name in sheet_names:
            ws = wb[sheet_name]

            for row in ws.iter_rows(min_row=1, max_row=50, max_col=15):
                for cell in row:
                    if cell.value is None:
                        continue

                    val = str(cell.value).strip()

                    # 宛先（御中）
                    if '御中' in val and not result["cover"].get("client"):
                        result["cover"]["client"] = val.replace('御中', '').replace('　', ' ').strip()

                    # 工事名
                    if '工事名' in val.replace(' ', '').replace('　', ''):
                        for c in range(cell.column + 1, min(cell.column + 6, 20)):
                            next_cell = ws.cell(row=cell.row, column=c)
                            if next_cell.value and len(str(next_cell.value).strip()) > 3:
                                result["cover"]["project_name"] = str(next_cell.value).strip()
                                break

                    # 工事場所・現場名
                    if '工事場所' in val or '現場名' in val:
                        for c in range(cell.column + 1, min(cell.column + 6, 20)):
                            next_cell = ws.cell(row=cell.row, column=c)
                            if next_cell.value:
                                result["cover"]["site_name"] = str(next_cell.value).strip()
                                break

                    # 住所
                    if '住所' in val and '現場' not in val:
                        for c in range(cell.column + 1, min(cell.column + 6, 20)):
                            next_cell = ws.cell(row=cell.row, column=c)
                            if next_cell.value:
                                result["cover"]["site_address"] = str(next_cell.value).strip()
                                break

                    # 見積日・日付
                    if ('見積日' in val or '日付' in val) and not result["cover"].get("quote_date"):
                        for c in range(cell.column + 1, min(cell.column + 6, 20)):
                            next_cell = ws.cell(row=cell.row, column=c)
                            if next_cell.value:
                                if isinstance(next_cell.value, datetime):
                                    result["cover"]["quote_date"] = next_cell.value.strftime('%Y-%m-%d')
                                else:
                                    result["cover"]["quote_date"] = str(next_cell.value).strip()
                                break

        # ============ 2. 内訳明細の解析 ============
        for sheet_name in sheet_names:
            # 条件書シートはスキップ
            if '条件' in sheet_name:
                continue

            ws = wb[sheet_name]

            for row in ws.iter_rows(min_row=1, max_row=300, values_only=False):
                row_values = [cell.value for cell in row]

                # 空行スキップ
                if not any(row_values):
                    continue

                first_val = str(row[0].value).strip() if row[0].value else ""

                # ヘッダー行・スキップ行
                if first_val in ['名称', '名　称', '品名', '項目', 'No.', 'NO.', '番号']:
                    continue
                if any(skip in first_val for skip in skip_keywords):
                    continue
                if not first_val:
                    continue

                # 明細行として解析
                try:
                    name = first_val
                    spec = ""
                    quantity = 0
                    unit = "式"
                    unit_price = 0
                    cost_price = 0
                    amount = 0

                    # 仕様（2列目）
                    if len(row) > 1 and row[1].value:
                        spec = str(row[1].value).strip()

                    # 数量（3列目）
                    if len(row) > 2 and row[2].value:
                        try:
                            quantity = float(row[2].value)
                        except (ValueError, TypeError):
                            pass

                    # 単位（4列目）
                    if len(row) > 3 and row[3].value:
                        unit_val = str(row[3].value).strip()
                        if any(u in unit_val for u in unit_keywords):
                            unit = unit_val

                    # 単価（5列目）
                    if len(row) > 4 and row[4].value:
                        try:
                            unit_price = float(row[4].value)
                        except (ValueError, TypeError):
                            pass

                    # 金額（6列目）
                    if len(row) > 5 and row[5].value:
                        try:
                            amount = float(row[5].value)
                        except (ValueError, TypeError):
                            pass

                    # 原価を探す（7列目以降）
                    for i in range(6, min(len(row), 10)):
                        if row[i].value:
                            try:
                                val = float(row[i].value)
                                if val > 0 and val != amount and val != unit_price:
                                    cost_price = val
                                    break
                            except (ValueError, TypeError):
                                pass

                    # 有効なデータのみ追加
                    if name and (quantity > 0 or amount > 0 or unit_price > 0):
                        # 金額がなければ計算
                        if amount == 0 and quantity > 0 and unit_price > 0:
                            amount = quantity * unit_price

                        result["items"].append({
                            "category": "",
                            "description": name,
                            "specification": spec,
                            "quantity": quantity if quantity > 0 else 1,
                            "unit": unit,
                            "unit_price": unit_price,
                            "cost_price": cost_price,
                            "amount": amount
                        })

                except Exception as e:
                    print(f"明細行パースエラー: {e}")
                    continue

        # ============ 3. 条件書の解析 ============
        for sheet_name in sheet_names:
            if '条件' not in sheet_name:
                continue

            ws = wb[sheet_name]

            for row in ws.iter_rows(min_row=2, max_row=100, max_col=5):
                category = str(row[0].value).strip() if row[0].value else ""
                content = str(row[1].value).strip() if len(row) > 1 and row[1].value else ""

                if category or content:
                    result["conditions"].append({
                        "category": category,
                        "content": content
                    })

        # ============ 4. 確認書（負担区分）の解析 ============
        for sheet_name in sheet_names:
            if '確認' not in sheet_name and '負担' not in sheet_name:
                continue

            ws = wb[sheet_name]

            for row in ws.iter_rows(min_row=2, max_row=50, max_col=5):
                item_name = str(row[0].value).strip() if row[0].value else ""

                if item_name and item_name not in ['項目', '名称']:
                    result["confirmation"]["items"].append({
                        "item": item_name,
                        "client": bool(row[1].value) if len(row) > 1 else False,
                        "company": bool(row[2].value) if len(row) > 2 else False,
                        "paid_supply": bool(row[3].value) if len(row) > 3 else False
                    })

        wb.close()

        # 成功レスポンス
        return {
            "success": True,
            "cover": result["cover"],
            "items": result["items"],
            "conditions": result["conditions"],
            "confirmation": result["confirmation"],
            "items_count": len(result["items"]),
            "conditions_count": len(result["conditions"]),
            "message": f"Excelを読み込みました（明細{len(result['items'])}件、条件{len(result['conditions'])}件）"
        }

    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excelファイルの解析に失敗しました: {str(e)}")
